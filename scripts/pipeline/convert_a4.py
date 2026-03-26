"""
convert_a4.py v2.1  —  A4 요약표 XLSX → merged.json (역방향)
- 파일·폴더·하위폴더 모두 처리 (rglob)
- 여러 시트 = 여러 사업 자동 처리
- 부처별/유형별 JSON 분리 저장 지원
실행:
  # 파일 하나
  python scripts/convert_a4.py output/A4요약_수정본.xlsx

  # 폴더 (하위폴더 포함 모든 xlsx)
  python scripts/convert_a4.py input/collect/

  # 여러 파일·폴더 동시
  python scripts/convert_a4.py input/부처A.xlsx input/collect/

  # 부처별로 분리 저장
  python scripts/convert_a4.py input/collect/ --split-by department

  # 특정 폴더에 분리 저장
  python scripts/convert_a4.py input/collect/ --split-by department --split-dir output/by_dept/

  # R&D/정보화/일반별 분리
  python scripts/convert_a4.py --split-by rnd
"""
import sys, re, json, yaml, datetime, logging, argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl --break-system-packages"); sys.exit(1)

ROOT    = Path(__file__).parent.parent.parent
ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


# ── 로거 ─────────────────────────────────────────────────────
def setup_logger(log_dir):
    log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    logger = logging.getLogger("kaib_a4")
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(log_dir / f"convert_a4_{ts}.log", encoding="utf-8")
    ch = logging.StreamHandler()
    fh.setLevel(logging.DEBUG); ch.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    fh.setFormatter(fmt); ch.setFormatter(fmt)
    logger.addHandler(fh); logger.addHandler(ch)
    return logger


# ── 헬퍼 ─────────────────────────────────────────────────────
def clean(v):
    if v is None: return None
    if isinstance(v, str): return ILLEGAL.sub("", v).strip() or None
    return v

def set_nested(obj, dotpath, value):
    parts = dotpath.split(".", 1)
    if len(parts) == 1:
        obj[dotpath] = value
    else:
        k, rest = parts
        if k not in obj or not isinstance(obj[k], dict):
            obj[k] = {}
        set_nested(obj[k], rest, value)

def get_nested(obj, dotpath):
    parts = dotpath.split(".", 1)
    if len(parts) == 1: return obj.get(dotpath)
    s = obj.get(parts[0])
    return get_nested(s, parts[1]) if isinstance(s, dict) else None

def deep_merge(base, override):
    result = base.copy()
    for k, v in override.items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = deep_merge(result[k], v)
        elif v is not None:
            result[k] = v
    return result


# ── Named Range 키 → JSON 필드 변환 ──────────────────────────
def nr_to_field(nr_name, sheet_prefix):
    """
    _1134_309_budget__2026_budget → budget.2026_budget
    _1134_309_project_period__start_year → project_period.start_year
    """
    prefix = f"_{sheet_prefix}_"
    if not nr_name.startswith(prefix):
        # 더 긴 prefix 시도
        parts = nr_name.split("_")
        # _AAAA_BBB_ 형태 prefix 제거 (최대 3개 언더스코어 그룹)
        for i in range(2, min(5, len(parts))):
            test_prefix = "_".join(parts[:i]) + "_"
            if nr_name.startswith("_" + test_prefix.lstrip("_")):
                field_part = nr_name[len("_" + test_prefix.lstrip("_")):]
                break
        else:
            return None
    else:
        field_part = nr_name[len(prefix):]

    # __ → .
    return field_part.replace("__", ".")


def parse_nr_ref(attr_text):
    """'1134-309'!$A$1 → (sname, row, col)"""
    m = re.match(r"'?([^'!]+)'?!\$([A-Z]+)\$(\d+)", attr_text)
    if not m: return None, None, None
    sname = m.group(1)
    col   = openpyxl.utils.column_index_from_string(m.group(2))
    row   = int(m.group(3))
    return sname, row, col


# ── 타입 복원 ─────────────────────────────────────────────────
BOOL_FIELDS   = {"is_rnd", "is_informatization"}
NUM_FIELDS    = {
    "budget.2024_settlement","budget.2025_original","budget.2025_supplementary",
    "budget.2026_request","budget.2026_budget","budget.change_amount","budget.change_rate",
    "total_cost.total","total_cost.government",
    "project_period.start_year","project_period.end_year","project_period.duration",
}
LIST_FIELDS   = {"ai_domains","ai_tech","rnd_stage","keywords"}

def restore_type(field, raw):
    if raw is None or raw == "": return None
    if field in BOOL_FIELDS:
        return str(raw).strip().upper() in ("O","TRUE","1","YES")
    if field in NUM_FIELDS:
        if isinstance(raw,(int,float)): return raw
        try:
            s = str(raw).replace(",","").strip()
            # 사업기간(년) 처리: "3년" → 3
            s = re.sub(r'[^0-9.\-]','', s)
            return float(s) if s and "." in s else int(s) if s else None
        except: return None
    if field in LIST_FIELDS:
        if isinstance(raw, list): return raw
        s = clean(str(raw))
        if not s: return []
        return [x.strip() for x in s.split(",") if x.strip()]
    # code 셀: "코드: 1134-309" 형태 정제
    if field == "code":
        s = clean(str(raw))
        if s and ":" in s:
            s = s.split(":",1)[1].strip()
        return s
    return clean(str(raw))


# ── 내역사업 행 스캔 ──────────────────────────────────────────
def scan_sub_projects(ws, logger):
    """
    "■ 내역사업" 섹션을 찾아 행 스캔
    반환: [{"name":..., "budget_2024":..., "budget_2025":..., "budget_2026":...}]
    """
    subs = []
    in_sub = False; in_data = False

    for row in ws.iter_rows():
        vals = [c.value for c in row]
        first = str(vals[0]).strip() if vals[0] else ""

        if "■ 내역사업" in first:
            in_sub  = True
            in_data = False
            continue
        if in_sub and first.startswith("■") and "내역사업" not in first:
            break
        if not in_sub: continue

        # 헤더 행 스킵
        if any(str(v) in ("내역사업명","내역사업","2024","2025","2026") for v in vals if v):
            in_data = True
            continue
        if not in_data: continue

        # 푸터·빈 행 종료
        if not any(v for v in vals): break
        if any(kw in str(vals[0] or "") for kw in ("출력일:","KAIB","자동 생성","※")): break

        name = clean(vals[0])
        if not name: continue

        def safe_num(v):
            if v is None: return None
            if isinstance(v,(int,float)): return float(v)
            try: return float(str(v).replace(",",""))
            except: return None

        subs.append({
            "name":        name,
            "budget_2024": safe_num(vals[2] if len(vals)>2 else None),
            "budget_2025": safe_num(vals[3] if len(vals)>3 else None),
            "budget_2026": safe_num(vals[4] if len(vals)>4 else None),
        })

    logger.debug(f"    내역사업 {len(subs)}건")
    return subs


# ── 사업연혁 행 스캔 ──────────────────────────────────────────
def scan_history(ws, logger):
    """■ 사업연혁 섹션 스캔"""
    hist = []
    in_sec = False; in_data = False

    for row in ws.iter_rows():
        vals = [c.value for c in row]
        first = str(vals[0]).strip() if vals[0] else ""

        if "■ 사업연혁" in first:
            in_sec = True; in_data = False; continue
        if in_sec and first.startswith("■") and "사업연혁" not in first: break
        if not in_sec: continue
        if any(str(v) in ("연도","주요 내용","내용") for v in vals if v):
            in_data = True; continue
        if not in_data: continue
        if not any(v for v in vals): break
        if any(kw in str(vals[0] or "") for kw in ("출력일:","KAIB","※")): break

        year = vals[0]
        desc = vals[1] if len(vals)>1 else None
        try: year = int(year)
        except: year = None
        if year and desc:
            hist.append({"year": year, "description": clean(str(desc))})

    logger.debug(f"    사업연혁 {len(hist)}건")
    return hist


# ── 시트 1장 → project dict ───────────────────────────────────
def extract_from_sheet(wb, ws, cfg, logger):
    sname = ws.title
    sheet_prefix = re.sub(r'[^A-Za-z0-9]','_', sname)[:8]

    # 이 시트에 속한 Named Range 수집
    sheet_nrs = {}
    for nr_name, dn in wb.defined_names.items():
        ref_sname, row, col = parse_nr_ref(dn.attr_text)
        if ref_sname != sname: continue
        field = nr_to_field(nr_name, sheet_prefix)
        if field: sheet_nrs[field] = (row, col)

    if not sheet_nrs:
        logger.warning(f"  [{sname}] Named Range 없음 — 스킵")
        return None

    # 기본 구조
    p = {
        "id": None, "name": None, "project_name": None, "code": None,
        "department": None, "division": None, "account_type": None,
        "field": None, "sector": None,
        "program": {"code": None, "name": None},
        "unit_project": {"code": None, "name": None},
        "detail_project": {"code": None, "name": None},
        "status": None, "support_type": None, "implementing_agency": None,
        "subsidy_rate": None, "loan_rate": None,
        "project_managers": [],
        "budget": {
            "2024_settlement": None, "2025_original": None,
            "2025_supplementary": None, "2026_request": None,
            "2026_budget": None, "change_amount": None, "change_rate": None,
        },
        "project_period": {"start_year": None, "end_year": None, "duration": None, "raw": None},
        "total_cost": {"total": None, "government": None, "raw": None},
        "sub_projects": [], "purpose": None, "description": None,
        "legal_basis": None, "is_rnd": False, "is_informatization": False,
        "keywords": [], "ai_domains": [], "ai_tech": None, "rnd_stage": None,
        "page_start": None, "page_end": None,
        "kpi": [], "overview": {}, "budget_calculation": [],
        "effectiveness": None, "execution_detail": {"method": None, "recipients": None, "subsidy_rate_detail": None},
        "yearly_budgets": {}, "history": [],
        "_source_file": getattr(wb, "_source_file", "unknown"),
        "_source_sheet": sname,
    }

    # Named Range → 값 추출
    for field, (row, col) in sheet_nrs.items():
        try:
            raw = ws.cell(row=row, column=col).value
            val = restore_type(field, raw)
            if val is not None:
                set_nested(p, field, val)
        except Exception as e:
            logger.debug(f"    [{sname}] {field} 추출 실패: {e}")

    # program.name 등에서 "[1100] 감사활동..." 형태 분리
    for obj_key in ("program", "unit_project", "detail_project"):
        obj = p.get(obj_key) or {}
        name_val = obj.get("name","") or ""
        m = re.match(r'\[(\w+)\]\s*(.*)', str(name_val))
        if m:
            obj["code"] = m.group(1)
            obj["name"] = m.group(2).strip()
        p[obj_key] = obj

    # 행 스캔: 내역사업
    p["sub_projects"] = scan_sub_projects(ws, logger)

    # 행 스캔: 사업연혁
    p["history"] = scan_history(ws, logger)

    # 자동 파생
    dept  = p.get("department","") or ""
    pname = p.get("project_name","") or ""
    if dept and pname:
        p["name"] = f"{dept}_{pname}"
    p["is_rnd"]            = "(R&D)" in pname
    p["is_informatization"] = "(정보화)" in pname

    # change_amount 재계산
    b6 = get_nested(p, "budget.2026_budget")
    b5 = get_nested(p, "budget.2025_original")
    if b6 is not None:
        if b5 is not None and get_nested(p,"budget.change_amount") is None:
            set_nested(p, "budget.change_amount", b6 - b5)
            if b5 != 0:
                set_nested(p, "budget.change_rate", round((b6-b5)/b5*100, 2))
        elif b5 is None and get_nested(p,"budget.change_amount") is None:
            set_nested(p, "budget.change_amount", b6)

    # yearly_budgets 갱신
    yb = p.get("yearly_budgets") or {}
    try:
        main_cfg_path = ROOT / "config" / "config.yaml"
        if main_cfg_path.exists():
            with open(main_cfg_path, encoding="utf-8") as f:
                mc = yaml.safe_load(f) or {}
            Y = mc.get("years", {})
            mapping = {
                Y.get("settlement"): get_nested(p,"budget.2024_settlement"),
                Y.get("original"):   get_nested(p,"budget.2025_original"),
                Y.get("budget"):     get_nested(p,"budget.2026_budget"),
            }
            for yr, val in mapping.items():
                if yr and val is not None:
                    yb[str(yr)] = val
    except Exception:
        pass
    p["yearly_budgets"] = yb

    logger.debug(f"  [{sname}] {p.get('project_name','?')} / {p.get('department','?')}")
    return p


# ── 파일 처리 ─────────────────────────────────────────────────
def convert_file(file_path, cfg, logger):
    file_path = Path(file_path)
    if not file_path.exists():
        logger.error(f"파일 없음: {file_path}"); return []

    logger.info(f"📂 {file_path.name}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    wb._source_file = file_path.name

    projects = []
    skipped  = 0
    for ws in wb.worksheets:
        p = extract_from_sheet(wb, ws, cfg, logger)
        if p is None:
            skipped += 1; continue
        if not p.get("project_name") or not p.get("department"):
            logger.warning(f"  [{ws.title}] 필수 필드 누락 — 스킵")
            skipped += 1; continue
        projects.append(p)

    logger.info(f"  → {len(projects)}건 추출 (스킵: {skipped}건)")
    return projects


# ── merged.json 저장 (upsert) ─────────────────────────────────
def save_merged(new_projects, mode, logger, split_by=None, split_out_dir=None):
    merged_path = ROOT / "output" / "merged.json"
    pk = "code"
    merged_path.parent.mkdir(parents=True, exist_ok=True)

    existing = []
    if mode != "overwrite" and merged_path.exists():
        try:
            with open(merged_path, encoding="utf-8") as f:
                data = json.load(f)
            existing = data.get("projects", data) if isinstance(data,dict) else data
            metadata = data.get("metadata", {}) if isinstance(data,dict) else {}
            logger.info(f"  기존: {len(existing)}건")
        except Exception as e:
            logger.warning(f"  기존 파일 읽기 실패: {e}")
            metadata = {}
    else:
        metadata = {}

    if mode == "upsert":
        exist_map = {p.get(pk): p for p in existing if p.get(pk)}
        for np in new_projects:
            pc = np.get(pk)
            if not pc: continue
            if pc in exist_map:
                old_yb = exist_map[pc].get("yearly_budgets") or {}
                new_yb = np.get("yearly_budgets") or {}
                exist_map[pc] = deep_merge(exist_map[pc], np)
                exist_map[pc]["yearly_budgets"] = {**old_yb, **new_yb}
            else:
                exist_map[pc] = np
        all_projects = list(exist_map.values())
    elif mode == "append":
        exist_keys = {p.get(pk) for p in existing if p.get(pk)}
        all_projects = existing + [p for p in new_projects if p.get(pk) not in exist_keys]
    else:
        all_projects = new_projects

    for i, p in enumerate(all_projects, 1):
        p["id"] = i

    total_2026 = sum((p.get("budget") or {}).get("2026_budget") or 0 for p in all_projects)
    total_2025 = sum((p.get("budget") or {}).get("2025_original") or 0 for p in all_projects)
    depts = {p.get("department") for p in all_projects if p.get("department")}

    merged = {
        "metadata": {
            **metadata,
            "total_projects":    len(all_projects),
            "total_departments": len(depts),
            "total_budget_2026": total_2026,
            "total_budget_2025": total_2025,
            "budget_change":     total_2026 - total_2025,
            "rnd_projects":      sum(1 for p in all_projects if p.get("is_rnd")),
            "info_projects":     sum(1 for p in all_projects if p.get("is_informatization")),
            "new_projects":      sum(1 for p in all_projects if p.get("status") == "신규"),
            "budget_mismatch_count": 0,
            "extraction_date":   datetime.date.today().isoformat(),
            "source":            "A4 요약표 역추출",
            "base_year":         cfg.get("base_year", 2026) if 'cfg' in locals() else 2026,
        },
        "projects": all_projects,
        "analysis": {
            "by_department": {}, "by_type": {}, "by_domain": {},
            "top_increases": [], "top_decreases": [],
            "duplicates": [], "duplicate_network": [],
            "keyword_clusters": [], "same_agency": [],
        },
    }

    with open(merged_path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    logger.info(f"✅ 저장: {merged_path} ({len(all_projects)}건, 모드:{mode})")

    # 분리 저장
    if split_by:
        split_dir = Path(split_out_dir) if split_out_dir else merged_path.parent / "split"
        split_dir.mkdir(parents=True, exist_ok=True)
        from collections import defaultdict
        groups = defaultdict(list)
        for p in all_projects:
            if split_by == "department":
                key = re.sub(r'[\\/*?:<>|]', '_', p.get("department","기타"))
            elif split_by == "status":
                key = p.get("status","기타")
            elif split_by == "rnd":
                key = "R&D" if p.get("is_rnd") else ("정보화" if p.get("is_informatization") else "일반")
            else:
                key = "전체"
            groups[key].append(p)

        for key, projs in groups.items():
            sub_merged = {**merged, "projects": projs,
                          "metadata": {**merged["metadata"],
                                        "total_projects": len(projs)}}
            sub_path = split_dir / f"{key}.json"
            with open(sub_path, "w", encoding="utf-8") as f:
                json.dump(sub_merged, f, ensure_ascii=False, indent=2)
            logger.info(f"  📄 분리: {sub_path.name} ({len(projs)}건)")

    return merged_path


# ── CLI ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="A4 XLSX → merged.json v2")
    parser.add_argument("targets", nargs="*",
        help="변환할 파일 또는 폴더 경로 (여러 개 가능, 하위 폴더 포함)")
    parser.add_argument("--overwrite", action="store_true",
        help="기존 merged.json 초기화 후 덮어쓰기")
    parser.add_argument("--append",    action="store_true",
        help="기존 merged.json에 추가 (중복 code 제외)")
    parser.add_argument("--split-by",
        choices=["department","status","rnd"], default=None,
        help="분리 저장 기준: department=부처별 | status=계속/신규별 | rnd=R&D/정보화/일반별")
    parser.add_argument("--split-dir", default=None,
        help="분리 저장 폴더 (기본: output/split/)")
    args = parser.parse_args()

    cfg = {}
    cfg_path = ROOT / "config" / "config_a4.yaml"
    if cfg_path.exists():
        with open(cfg_path, encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}

    log_dir = ROOT / "logs"
    try:
        mc_path = ROOT / "config" / "config.yaml"
        if mc_path.exists():
            with open(mc_path, encoding="utf-8") as f:
                mc = yaml.safe_load(f) or {}
            log_dir = ROOT / mc.get("paths",{}).get("logs","logs")
    except Exception: pass

    logger = setup_logger(log_dir)

    # 모드 결정
    if args.overwrite:  mode = "overwrite"
    elif args.append:   mode = "append"
    else:               mode = "upsert"

    # 대상 파일 수집
    if args.targets:
        paths = []
        for t in args.targets:
            t = Path(t)
            if t.is_dir():
                # 하위 폴더 포함 모든 xlsx 탐색 (rglob)
                found = list(t.rglob("*.xlsx"))
                logger.info(f"  폴더 탐색: {t} → {len(found)}개 파일")
                paths += found
            else:
                paths.append(t)
    else:
        # 인자 없으면 input/ 폴더 전체 탐색
        input_dir = ROOT / "input"
        paths = list(input_dir.rglob("*.xlsx"))
        if not paths:
            paths = list((ROOT/"output").glob("A4요약*.xlsx"))

    if not paths:
        logger.warning("A4 요약표 파일 없음. 경로를 직접 지정하세요.")
        return

    logger.info(f"▶ {len(paths)}개 파일 변환 시작")
    all_projects = []
    for p in paths:
        projects = convert_file(p, cfg, logger)
        all_projects.extend(projects)

    if all_projects:
        merged_path = save_merged(all_projects, mode, logger,
                                    split_by=args.split_by,
                                    split_out_dir=args.split_dir)
        print(f"\n📌 다음: python scripts/build_analysis.py")
        print(f"   또는: cp {merged_path} ../KAIB2026/data/budget_db.json")
    else:
        logger.warning("추출된 사업 없음")


if __name__ == "__main__":
    main()