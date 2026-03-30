"""
convert.py v2.0  —  XLSX → merged.json
설계 확정 반영:
  - code = PK (upsert 지원)
  - yearly_budgets 누적
  - 전체 필드 정합 (budget_db.json 기준)
  - kpi/history/yearly_budgets 시트 처리
  - defaults 자동 채움
실행: python scripts/convert.py [파일/폴더] [--overwrite] [--watch]
"""
import sys, os, json, yaml, re, datetime, logging, argparse
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl --break-system-packages"); sys.exit(1)

ROOT = Path(__file__).parent.parent.parent
import sys as _sys; _sys.path.insert(0, str(ROOT/"scripts"))
from _years import get_years as _get_years_module

def get_years(cfg=None):
    return _get_years_module(cfg if cfg else ROOT/"config"/"config.yaml")
ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


# ── 로거 ──────────────────────────────────────────────────────
def setup_logger(log_dir):
    log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    logger = logging.getLogger("kaib")
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(log_dir / f"convert_{ts}.log", encoding="utf-8")
    ch = logging.StreamHandler()
    fh.setLevel(logging.DEBUG); ch.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    fh.setFormatter(fmt); ch.setFormatter(fmt)
    logger.addHandler(fh); logger.addHandler(ch)
    return logger


# ── config 로드 ───────────────────────────────────────────────
def load_config():
    p = ROOT / "config" / "config.yaml"
    if not p.exists(): raise FileNotFoundError(f"config.yaml 없음: {p}")
    with open(p, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    # _years 기반으로 column_mapping 동적 교체
    Y = get_years(cfg)
    xlsx = cfg.setdefault("xlsx", {})
    
    # 플레이스홀더 치환 지원 ({base}, {prev}, {settlement})
    subs = {
        "{base}":       str(Y["budget"]),
        "{prev}":       str(Y["original"]),
        "{settlement}": str(Y["settlement"]),
    }

    def resolve(v):
        if isinstance(v, str):
            for k, s in subs.items(): v = v.replace(k, s)
        elif isinstance(v, list):
            return [resolve(x) for x in v]
        elif isinstance(v, dict):
            return {rk: resolve(rv) for rk, rv in v.items()}
        return v

    # 1. xlsx 매핑 및 설정 치환
    if "column_mapping" in xlsx:
        xlsx["column_mapping"] = resolve(xlsx["column_mapping"])
    if "sub_projects_mapping" in xlsx:
        xlsx["sub_projects_mapping"] = resolve(xlsx["sub_projects_mapping"])
    
    # 2. schema 컬럼명 치환
    if "schema" in cfg:
        cfg["schema"] = resolve(cfg["schema"])

    # _budget_* 내부 키 → 실제 연도 레이블 키로 교체 (엑셀 컬럼 매칭용)
    col_map = xlsx.get("column_mapping", {})
    internal_to_label = {
        "_budget_settlement":    Y["label_settlement"],
        "_budget_original":      Y["label_original"],
        "_budget_supplementary": Y["label_supplementary"],
        "_budget_request":       Y["label_request"],
        "_budget_confirmed":     Y["label_budget"],
    }
    for internal_key, label in internal_to_label.items():
        if internal_key in col_map:
            val = col_map.pop(internal_key)
            col_map[label]         = val
            col_map[f"★{label}"] = val

    # stale_keys 제거
    stale_keys = [k for k in list(col_map.keys())
                  if any(x in k for x in ["결산","본예산","추경","요구","정부안","확정"])
                  and not k.startswith("_")
                  and not any(str(y) in k for y in [Y["settlement"],Y["original"],Y["budget"]])]
    for k in stale_keys:
        col_map.pop(k, None)

    # 내역사업 _sub_year* → 실제 연도 레이블로 교체
    sub_map = xlsx.get("sub_projects_mapping", {})
    sub_internal = {
        "_sub_year0": Y["label_sub"][0],
        "_sub_year1": Y["label_sub"][1],
        "_sub_year2": Y["label_sub"][2],
    }
    for internal_key, label in sub_internal.items():
        if internal_key in sub_map:
            val = sub_map.pop(internal_key)
            sub_map[label]         = val
            sub_map[f"★{label}"] = val

    # 3. derived_fields도 years 기반으로 동적 생성
    derived = cfg.setdefault("derived_fields", {})
    b_key = f"budget.{Y['budget']}_budget"
    o_key = f"budget.{Y['original']}_original"
    
    derived["budget.change_amount"] = {
        "formula": f"{b_key} - {o_key}",
        "condition": f"{b_key} is not null and {o_key} is not null"
    }
    derived["budget.change_rate"] = {
        "formula": f"(budget.change_amount / {o_key}) * 100",
        "condition": f"{o_key} is not null and {o_key} != 0"
    }

    return cfg


# ── 헬퍼 ──────────────────────────────────────────────────────
def clean(v):
    if v is None: return None
    if isinstance(v, str): return ILLEGAL.sub("", v).strip() or None
    return v

def set_nested(obj, dotpath, value):
    parts = dotpath.split(".", 1)
    if len(parts) == 1:
        obj[dotpath] = value
    else:
        k, rest = parts
        if k not in obj or not isinstance(obj[k], dict):
            obj[k] = {}
        set_nested(obj[k], rest, value)

def get_nested(obj, dotpath):
    parts = dotpath.split(".", 1)
    if len(parts) == 1: return obj.get(dotpath)
    s = obj.get(parts[0])
    return get_nested(s, parts[1]) if isinstance(s, dict) else None

def deep_merge(base, override):
    """base dict에 override를 재귀 병합 (override 우선)"""
    result = base.copy()
    for k, v in override.items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = deep_merge(result[k], v)
        else:
            result[k] = v
    return result


# ── 타입 변환 ─────────────────────────────────────────────────
def apply_converter(value, conv_name, converters_cfg, logger):
    if value is None or value == "":
        cfg = converters_cfg.get(conv_name, {})
        t = cfg.get("type","str")
        if t == "list": return cfg.get("on_error", [])
        return cfg.get("on_error", None)

    cfg = converters_cfg.get(conv_name, {})
    t = cfg.get("type", "str")
    try:
        if t == "int":
            return int(float(str(value).replace(",","").strip()))
        elif t == "float":
            v = str(value).replace(",","").strip()
            return float(v) if v else cfg.get("on_error", None)
        elif t == "list":
            sep = cfg.get("separator", ",")
            items = [x.strip() for x in str(value).split(sep)]
            return [x for x in items if x]
        elif t == "text":
            # 줄바꿈 보존, 앞뒤 공백 제거
            return ILLEGAL.sub("", str(value)).strip() or None
        else:
            return clean(str(value))
    except Exception as e:
        logger.debug(f"  변환 실패 ({conv_name}): '{value}' → {e}")
        return cfg.get("on_error", None)


# ── 파생 필드 계산 ────────────────────────────────────────────
def apply_derived(project, derived_cfg, logger):
    for field, rule in derived_cfg.items():
        formula = rule.get("formula","")
        condition = rule.get("condition","")
        try:
            if condition and not _eval_cond(project, condition):
                continue
            val = _eval_formula(project, formula)
            if val is not None:
                set_nested(project, field, val)
        except Exception as e:
            logger.debug(f"  파생 실패 ({field}): {e}")

def _eval_formula(p, formula):
    if "{" in formula:
        flat = _flatten(p)
        try: return formula.format(**flat)
        except KeyError: return None
    m = re.match(r"contains\((.+?),\s*'(.+?)'\)", formula)
    if m:
        return (m.group(2) in str(get_nested(p, m.group(1)) or ""))
    m2 = re.match(r"str\((.+?)\)\s*\+\s*'(.+?)'", formula)
    if m2:
        v = _eval_formula(p, m2.group(1))
        return str(v) + m2.group(2) if v is not None else None
    expr = formula
    for match in re.finditer(r"[a-zA-Z_][a-zA-Z0-9_.]*", formula):
        fn = match.group()
        val = get_nested(p, fn)
        if val is None: return None
        expr = expr.replace(fn, str(val))
    try: return eval(expr)
    except: return None

def _eval_cond(p, cond):
    if " and " in cond:
        return all(_eval_cond(p, c.strip()) for c in cond.split(" and "))
    m = re.match(r"(.+?)\s+is\s+not\s+null", cond)
    if m: return get_nested(p, m.group(1).strip()) is not None
    m2 = re.match(r"(.+?)\s*!=\s*(.+)", cond)
    if m2:
        val = get_nested(p, m2.group(1).strip())
        try: return val != float(m2.group(2).strip())
        except: return val != m2.group(2).strip()
    return True

def _flatten(d, prefix=""):
    r = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        simple = k
        if isinstance(v, dict): r.update(_flatten(v, key))
        else: r[simple] = v if v is not None else ""
    return r


# ── yearly_budgets 갱신 ───────────────────────────────────────
def update_yearly_budgets(project, years_cfg):
    """budget 슬라이딩 값을 yearly_budgets에 누적"""
    yb = project.get("yearly_budgets") or {}
    b  = project.get("budget") or {}
    mapping = {
        years_cfg.get("settlement"): b.get("2024_settlement"),
        years_cfg.get("original"):   b.get("2025_original"),
        years_cfg.get("budget"):     b.get("2026_budget"),
    }
    for yr, val in mapping.items():
        if yr and val is not None:
            yb[str(yr)] = val
    project["yearly_budgets"] = yb


# ── 빈 project 구조 생성 ──────────────────────────────────────
def make_empty_project(defaults):
    import copy
    base = {
        "id": None, "name": None, "project_name": None, "code": None,
        "department": None, "division": None, "account_type": None,
        "field": None, "sector": None,
        "program": {"code": None, "name": None},
        "unit_project": {"code": None, "name": None},
        "detail_project": {"code": None, "name": None},
        "status": None, "support_type": None, "implementing_agency": None,
        "subsidy_rate": None, "loan_rate": None,
        "project_managers": [],
        "budget": {
            "2024_settlement": None, "2025_original": None,
            "2025_supplementary": None, "2026_request": None,
            "2026_budget": None, "change_amount": None, "change_rate": None,
        },
        "project_period": {"start_year": None, "end_year": None, "duration": None, "raw": None},
        "total_cost": {"total": None, "government": None, "raw": None},
        "sub_projects": [], "purpose": None, "description": None,
        "legal_basis": None, "is_rnd": False, "is_informatization": False,
        "keywords": [], "page_start": None, "page_end": None,
        "kpi": [], "overview": {}, "budget_calculation": [],
        "effectiveness": "(작성예정)",
        "execution_detail": {"method": None, "recipients": None, "subsidy_rate_detail": None},
        "yearly_budgets": {}, "history": [], "ai_domains": [],
        "ai_tech": None, "rnd_stage": None,
    }
    # defaults 적용
    if defaults:
        for k, v in defaults.items():
            if k in base and base[k] in (None, [], {}, False):
                base[k] = copy.deepcopy(v)
    return base


# ── 헤더 인덱스 빌드 ──────────────────────────────────────────
def build_header_idx(ws, header_row):
    idx = {}
    for ci in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=ci).value
        if val:
            raw = str(val).strip()
            clean_h = raw.replace("★", "").strip()
            idx[raw] = ci
            idx[clean_h] = ci
    return idx


# ── 사업목록 시트 읽기 ────────────────────────────────────────
def read_projects_sheet(wb, cfg, logger):
    xlsx = cfg["xlsx"]
    conv = cfg.get("converters", {})
    defs = cfg.get("defaults", {})
    derived = cfg.get("derived_fields", {})
    years = cfg.get("years", {})
    valid = cfg.get("validation", {})
    col_map = xlsx.get("column_mapping", {})

    sname = xlsx["data_sheet"]
    if sname not in wb.sheetnames:
        logger.error(f"시트 없음: {sname}"); return []
    ws = wb[sname]

    header_idx = build_header_idx(ws, xlsx["header_row"])
    logger.debug(f"  헤더 감지: {list(header_idx.keys())[:8]}...")

    # 컬럼 매핑 빌드
    field_map = []
    for hkey, jspec in col_map.items():
        ci = header_idx.get(hkey) or header_idx.get(hkey.replace("★","").strip())
        if ci is None:
            logger.debug(f"  헤더 미발견 (무시): '{hkey}'")
            continue
        if "|" in jspec:
            jf, cv_name = jspec.split("|", 1)
        else:
            jf, cv_name = jspec, None
        field_map.append((ci, jf.strip(), cv_name))

    projects = []
    stop = xlsx.get("stop_on_empty", True)

    for row in range(xlsx["data_start_row"], ws.max_row + 1):
        first3 = [ws.cell(row=row, column=c).value for c in range(1, 4)]
        if stop and all(v is None or str(v).strip() == "" for v in first3):
            logger.debug(f"  빈 행 → {row}행에서 중단")
            break

        p = make_empty_project(defs)

        for ci, jf, cv_name in field_map:
            raw = ws.cell(row=row, column=ci).value
            if cv_name:
                val = apply_converter(raw, cv_name, conv, logger)
            else:
                val = clean(str(raw)) if raw is not None else None
            if val is not None:
                set_nested(p, jf, val)

        # 파생 필드 계산
        apply_derived(p, derived, logger)

        # 원본 호환: 2025_original=None이면 change_amount=2026_budget (budget_db 호환)
        b = p.get('budget') or {}
        if (b.get('2026_budget') is not None
                and b.get('2025_original') is None
                and b.get('change_amount') is None):
            b['change_amount'] = b['2026_budget']
            b['change_rate']   = None

        # yearly_budgets 갱신
        update_yearly_budgets(p, years)

        # 메타
        p["_source_file"] = getattr(wb, "_source_file", "unknown")
        p["_source_row"]  = row

        # 유효성 검사
        errors = []
        for req in valid.get("required_fields", []):
            if not get_nested(p, req):
                errors.append(f"필수 누락: {req}")
        for e in errors:
            logger.warning(f"  [행{row}] ❌ {e}")
        if errors and valid.get("strict_mode", False):
            raise ValueError(f"유효성 실패 행{row}: {errors}")

        projects.append(p)
        logger.debug(f"  ✓ 행{row}: {p.get('project_name')} ({p.get('department')})")

    logger.info(f"  → 사업목록: {len(projects)}건")
    return projects


# ── 내역사업 시트 읽기 ────────────────────────────────────────
def read_sub_sheet(wb, cfg, projects, logger):
    xlsx = cfg["xlsx"]
    sname = xlsx.get("sub_projects_sheet")
    if not sname or sname not in wb.sheetnames: return
    ws = wb[sname]
    mapping = xlsx.get("sub_projects_mapping", {})
    conv = cfg.get("converters", {})

    header_idx = build_header_idx(ws, 2)
    field_map = []
    for hk, jspec in mapping.items():
        ci = header_idx.get(hk) or header_idx.get(hk.replace("★","").strip())
        if ci is None: continue
        jf, cv_name = (jspec.split("|",1) + [None])[:2]
        field_map.append((ci, jf.strip(), cv_name))

    code_to_idx = {p.get("code"): i for i, p in enumerate(projects) if p.get("code")}

    for row in range(3, ws.max_row + 1):
        if all(ws.cell(row=row, column=c).value is None for c in range(1, 4)): break
        sub = {}
        for ci, jf, cv_name in field_map:
            raw = ws.cell(row=row, column=ci).value
            val = apply_converter(raw, cv_name, conv, logger) if cv_name else clean(str(raw)) if raw else None
            sub[jf] = val
        pc = sub.pop("parent_code", None)
        if pc and pc in code_to_idx:
            projects[code_to_idx[pc]]["sub_projects"].append(sub)

    logger.info("  → 내역사업 시트 처리 완료")


# ── 사업관리자 시트 읽기 ──────────────────────────────────────
def read_managers_sheet(wb, cfg, projects, logger):
    xlsx = cfg["xlsx"]
    sname = xlsx.get("managers_sheet")
    if not sname or sname not in wb.sheetnames: return
    ws = wb[sname]
    mapping = xlsx.get("managers_mapping", {})
    header_idx = build_header_idx(ws, 2)
    field_map = []
    for hk, jf in mapping.items():
        ci = header_idx.get(hk) or header_idx.get(hk.replace("★","").strip())
        if ci: field_map.append((ci, jf.strip()))

    code_to_idx = {p.get("code"): i for i, p in enumerate(projects) if p.get("code")}

    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is None: break
        mgr = {}
        for ci, jf in field_map:
            raw = ws.cell(row=row, column=ci).value
            mgr[jf] = clean(str(raw)) if raw else None
        pc = mgr.pop("parent_code", None)
        if pc and pc in code_to_idx:
            projects[code_to_idx[pc]]["project_managers"].append(mgr)

    logger.info("  → 사업관리자 시트 처리 완료")


# ── 사업연혁 시트 읽기 ────────────────────────────────────────
def read_history_sheet(wb, cfg, projects, logger):
    xlsx = cfg["xlsx"]
    sname = xlsx.get("history_sheet")
    if not sname or sname not in wb.sheetnames: return
    ws = wb[sname]
    mapping = xlsx.get("history_mapping", {})
    conv = cfg.get("converters", {})
    header_idx = build_header_idx(ws, 2)
    field_map = []
    for hk, jspec in mapping.items():
        ci = header_idx.get(hk) or header_idx.get(hk.replace("★","").strip())
        if ci is None: continue
        jf, cv_name = (jspec.split("|",1) + [None])[:2]
        field_map.append((ci, jf.strip(), cv_name))

    code_to_idx = {p.get("code"): i for i, p in enumerate(projects) if p.get("code")}

    for row in range(3, ws.max_row + 1):
        if all(ws.cell(row=row, column=c).value is None for c in range(1, 3)): break
        hist = {}
        for ci, jf, cv_name in field_map:
            raw = ws.cell(row=row, column=ci).value
            val = apply_converter(raw, cv_name, conv, logger) if cv_name else clean(str(raw)) if raw else None
            hist[jf] = val
        pc = hist.pop("parent_code", None)
        if pc and pc in code_to_idx:
            projects[code_to_idx[pc]]["history"].append(hist)

    logger.info("  → 사업연혁 시트 처리 완료")


# ── 연도별예산 시트 읽기 ──────────────────────────────────────
def read_yearly_budget_sheet(wb, cfg, projects, logger):
    xlsx = cfg["xlsx"]
    sname = xlsx.get("yearly_budget_sheet")
    if not sname or sname not in wb.sheetnames: return
    ws = wb[sname]
    header_idx = build_header_idx(ws, 2)
    code_to_idx = {p.get("code"): i for i, p in enumerate(projects) if p.get("code")}

    # 연도 컬럼 동적 감지 (YYYY 형태)
    year_cols = {}
    for hk, ci in header_idx.items():
        try:
            yr = int(hk)
            if 2000 <= yr <= 2100:
                year_cols[str(yr)] = ci
        except (ValueError, TypeError):
            pass
    code_col = header_idx.get("상위사업코드") or header_idx.get("사업코드")
    if not code_col: return

    for row in range(3, ws.max_row + 1):
        pc_val = ws.cell(row=row, column=code_col).value
        if not pc_val: break
        pc = clean(str(pc_val))
        if pc and pc in code_to_idx:
            yb = projects[code_to_idx[pc]].get("yearly_budgets") or {}
            for yr_str, ci in year_cols.items():
                v = ws.cell(row=row, column=ci).value
                if v is not None:
                    try: yb[yr_str] = float(str(v).replace(",",""))
                    except: pass
            projects[code_to_idx[pc]]["yearly_budgets"] = yb

    logger.info("  → 연도별예산 시트 처리 완료")


# ── 파일 변환 ─────────────────────────────────────────────────
def convert_file(file_path, cfg, logger):
    file_path = Path(file_path)
    if not file_path.exists():
        logger.error(f"파일 없음: {file_path}"); return []
    logger.info(f"📂 {file_path.name}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    wb._source_file = file_path.name

    projects = read_projects_sheet(wb, cfg, logger)
    if not projects: return []

    read_sub_sheet(wb, cfg, projects, logger)
    read_managers_sheet(wb, cfg, projects, logger)
    read_history_sheet(wb, cfg, projects, logger)
    read_yearly_budget_sheet(wb, cfg, projects, logger)

    # 개별 JSON 저장
    out_cfg = cfg.get("output", {})
    if out_cfg.get("save_individual", False):
        ind_dir = ROOT / out_cfg.get("individual_dir", "output/individual")
        ind_dir.mkdir(parents=True, exist_ok=True)
        ind_path = ind_dir / (file_path.stem + ".json")
        with open(ind_path, "w", encoding="utf-8") as f:
            json.dump(projects, f, ensure_ascii=False, indent=2)
        logger.info(f"  💾 개별 저장: {ind_path.name} ({len(projects)}건)")

    return projects


def read_json_file(file_path, cfg, logger):
    """과거 JSON 파일을 입력받을 때 연도 프로퍼티를 config의 base_year로 자동 치환하는 마이그레이터"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        logger.error(f"JSON 파싱 실패 {file_path}: {e}")
        return []

    p_list = data.get("projects", data) if isinstance(data, dict) else data
    if not isinstance(p_list, list): return []
    
    Y = get_years(cfg)
    by = cfg.get("years", {}).get("base_year", 2026)
    
    valid_projects = []
    for p in p_list:
        if not isinstance(p, dict): continue
        if not p.get("code") or not p.get("project_name"): continue
        
        # Budget Dictionary Key Migration
        b = p.get("budget", {})
        if isinstance(b, dict):
            new_b = {}
            for k, v in b.items():
                if k in ["change_amount", "change_rate"]: 
                    new_b[k] = v
                    continue
                m = re.match(r"\d{4}_(.*)", k)
                if m:
                    suffix = m.group(1)
                    if suffix == "settlement": new_yr = by - 2
                    elif suffix == "original": new_yr = by - 1
                    elif suffix == "supplementary": new_yr = by - 1
                    elif suffix in ["request", "budget"]: new_yr = by
                    else: new_yr = by
                    new_b[f"{new_yr}_{suffix}"] = v
                else:
                    new_b[k] = v
            p["budget"] = new_b
            
        apply_derived(p, cfg.get("derived_fields", {}), logger)
        update_yearly_budgets(p, cfg.get("years", {}))
        valid_projects.append(p)
        
    logger.info(f"📂 {file_path.name} (JSON 마이그레이션 적용 완료: {len(valid_projects)}건)")
    return valid_projects


# ── merged.json 저장 (upsert 지원) ────────────────────────────
def save_merged(new_projects, cfg, logger):
    out_cfg   = cfg.get("output", {})
    merged_path = ROOT / cfg["paths"]["merged"]
    mode = out_cfg.get("merge_mode", "upsert")
    pk   = cfg.get("pk_field", "code")
    excl = set(cfg.get("exclude_codes", []) or [])
    merged_path.parent.mkdir(parents=True, exist_ok=True)

    existing = []
    if mode != "overwrite" and merged_path.exists():
        try:
            with open(merged_path, encoding="utf-8") as f:
                data = json.load(f)
            existing = data.get("projects", data) if isinstance(data, dict) else data
            logger.info(f"  기존 merged.json: {len(existing)}건")
        except Exception as e:
            logger.warning(f"  기존 파일 읽기 실패: {e}")

    # upsert: code 기준으로 기존 레코드 교체
    if mode == "upsert":
        existing_map = {p.get(pk): p for p in existing if p.get(pk)}
        for np in new_projects:
            pc = np.get(pk)
            if not pc: continue
            if pc in existing_map:
                # yearly_budgets 누적 병합
                old_yb = existing_map[pc].get("yearly_budgets") or {}
                new_yb = np.get("yearly_budgets") or {}
                merged_yb = {**old_yb, **new_yb}  # 새 값 우선
                existing_map[pc] = deep_merge(existing_map[pc], np)
                existing_map[pc]["yearly_budgets"] = merged_yb
                logger.debug(f"  upsert: {pc}")
            else:
                existing_map[pc] = np
        all_projects = list(existing_map.values())
    elif mode == "append":
        exist_keys = {p.get(pk) for p in existing if p.get(pk)}
        all_projects = existing + [p for p in new_projects if p.get(pk) not in exist_keys]
    else:  # overwrite
        all_projects = new_projects

    # 제외 코드 필터
    all_projects = [p for p in all_projects if p.get(pk) not in excl]

    # 고정값(Stable) 해시 ID 재발번 (AI 의존성 끊김 방지)
    import hashlib
    for p in all_projects:
        raw_str = f"{p.get('department', '')}_{p.get('project_name', '')}_{p.get('code', '')}"
        stable_hash = hashlib.md5(raw_str.encode('utf-8')).hexdigest()[:8].upper()
        p["id"] = f"PRJ-{stable_hash}"

    # metadata 계산
    total_2026 = sum((p.get("budget") or {}).get("2026_budget") or 0 for p in all_projects)
    total_2025 = sum((p.get("budget") or {}).get("2025_original") or 0 for p in all_projects)
    depts = {p.get("department") for p in all_projects if p.get("department")}

    merged = {
        "metadata": {
            "total_projects":    len(all_projects),
            "total_departments": len(depts),
            "total_budget_2026": total_2026,
            "total_budget_2025": total_2025,
            "budget_change":     total_2026 - total_2025,
            "rnd_projects":      sum(1 for p in all_projects if p.get("is_rnd")),
            "info_projects":     sum(1 for p in all_projects if p.get("is_informatization")),
            "new_projects":      sum(1 for p in all_projects if p.get("status") == "신규"),
            "budget_mismatch_count": 0,
            "extraction_date":   datetime.date.today().isoformat(),
            "source":            cfg.get("metadata", {}).get("source", "KAIB 파이프라인"),
            "base_year":         cfg.get("base_year", 2026),
        },
        "projects": all_projects,
        "analysis": {
            "by_department": {}, "by_type": {}, "by_domain": {},
            "top_increases": [], "top_decreases": [],
            "duplicates": [], "duplicate_network": [],
            "keyword_clusters": [], "same_agency": [],
        },
    }

    with open(merged_path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    logger.info(f"✅ merged.json 저장: {merged_path}")
    logger.info(f"   총 {len(all_projects)}건 (모드: {mode})")
    return merged_path


# ── watch 모드 ────────────────────────────────────────────────
def watch_mode(cfg, logger):
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
    except ImportError:
        logger.error("watchdog 미설치: pip install watchdog --break-system-packages"); sys.exit(1)

    input_dir = ROOT / cfg["paths"]["input"]
    input_dir.mkdir(parents=True, exist_ok=True)

    class Handler(FileSystemEventHandler):
        def on_created(self, event):
            if not event.is_directory:
                p = Path(event.src_path)
                if p.suffix.lower() in (".xlsx", ".xlsm"):
                    logger.info(f"🔍 감지: {p.name}")
                    import time; time.sleep(1)
                    projects = convert_file(p, cfg, logger)
                    if projects:
                        save_merged(projects, cfg, logger)
                        after = cfg["pipeline"].get("after_convert", "keep")
                        if after == "archive":
                            arch = ROOT / cfg["pipeline"].get("archive_dir", "input/processed")
                            arch.mkdir(parents=True, exist_ok=True)
                            p.rename(arch / p.name)
                        elif after == "delete":
                            p.unlink()

    obs = Observer()
    obs.schedule(Handler(), str(input_dir), recursive=False)
    obs.start()
    logger.info(f"👁  watch 모드: {input_dir} 감시 중 (Ctrl+C 종료)")
    try:
        import time
        while True: time.sleep(cfg["pipeline"].get("watch_interval", 3))
    except KeyboardInterrupt:
        obs.stop()
    obs.join()


# ── CLI ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="XLSX → merged.json (v2)")
    parser.add_argument("targets", nargs="*")
    parser.add_argument("--watch",     action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    cfg = load_config()
    if args.overwrite:
        cfg["output"]["merge_mode"] = "overwrite"

    log_dir = ROOT / cfg["paths"].get("logs", "logs")
    logger = setup_logger(log_dir)

    if args.watch or cfg["pipeline"].get("run_mode") == "watch":
        watch_mode(cfg, logger); return

    if args.targets:
        paths = []
        for t in args.targets:
            t = Path(t)
            if t.is_dir(): paths += list(t.glob("*.xlsx")) + list(t.glob("*.xlsm"))
            else: paths.append(t)
    else:
        input_dir = ROOT / cfg["paths"]["input"]
        paths = list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xlsm")) + list(input_dir.glob("*.json"))
        if not paths:
            logger.warning(f"input/ 에 변환할 파일(xlsx, json)이 없습니다: {input_dir}"); return

    logger.info(f"▶ 변환 시작: {len(paths)}개 파일")
    all_projects = []
    for p in paths:
        if p.suffix.lower() == ".json":
            projects = read_json_file(p, cfg, logger)
        else:
            projects = convert_file(p, cfg, logger)
            
        if projects:
            all_projects.extend(projects)

    if all_projects:
        merged_path = save_merged(all_projects, cfg, logger)
        print(f"\n📌 다음: cp {merged_path} ../KAIB2026/data/budget_db.json")
    else:
        logger.warning("변환된 사업 없음")


if __name__ == "__main__":
    main()
