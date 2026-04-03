"""
export_xlsx.py v2.0  —  merged.json → 총괄 XLSX (검토용)
- 전체 필드 정합 반영
- config_export.yaml 기준 컬럼 선택
실행: python scripts/export_xlsx.py [입력.json] [--out 출력.xlsx]
"""
import sys, re, json, yaml, datetime, argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("pip install openpyxl --break-system-packages"); sys.exit(1)

ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))
from config import path_config
from scripts.pipeline._years import get_years as _gy

def get_years(cfg=None):
    return _gy(cfg if cfg else ROOT/"config"/"config.yaml")

ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def load_cfg(p):
    with open(p, encoding="utf-8") as f: return yaml.safe_load(f)

def gn(obj, path):
    if not isinstance(obj, dict): return None
    parts = path.split(".", 1)
    if len(parts) == 1: return obj.get(path)
    return gn(obj.get(parts[0]), parts[1])

def fv(val):
    if val is None: return ""
    if isinstance(val, bool): return "O" if val else ""
    if isinstance(val, list): return ", ".join(ILLEGAL.sub("", str(v)) for v in val)
    if isinstance(val, str):  return ILLEGAL.sub("", val)
    return val

def hfl(c): return PatternFill("solid", fgColor=c.lstrip("#"))
def tb(c="CCCCCC"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def hb():
    s = Side(style="thin", color="CCCCCC")
    b = Side(style="medium", color="4A90D9")
    return Border(left=s, right=s, top=s, bottom=b)


# ── 요약 시트 ────────────────────────────────────────────────
def build_summary(ws, meta, cfg):
    fn  = cfg.get("style",{}).get("font_name","맑은 고딕")
    hbg = cfg.get("style",{}).get("header_bg","1F3864")
    ws.title = cfg.get("sheet_summary",{}).get("name","요약")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "AI 재정사업 총괄 요약"
    c.font = Font(name=fn, size=13, bold=True, color="FFFFFF")
    c.fill = hfl(hbg); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    Y = get_years()
    rows = [
        ("총 사업 수",         f"{meta.get('total_projects',0):,}개"),
        ("소관 부처 수",       f"{meta.get('total_departments',0):,}개"),
        ("R&D 사업",           f"{meta.get('rnd_projects',0):,}개"),
        ("정보화 사업",        f"{meta.get('info_projects',0):,}개"),
        ("신규 사업",          f"{meta.get('new_projects',0):,}개"),
        (f"{Y['original']} 본예산 합계",   f"{meta.get('total_budget_' + Y['original'], 0):,.0f} 백만원"),
        (f"{Y['budget']} 확정예산 합계", f"{meta.get('total_budget_' + Y['budget'], 0):,.0f} 백만원"),
        ("전년 대비 증감",     f"{meta.get('budget_change',0):+,.0f} 백만원"),
        ("데이터 추출일",      meta.get("extraction_date","")),
        ("원본 자료",          meta.get("source","")),
    ]
    for i,(lbl,val) in enumerate(rows, 2):
        cl = ws.cell(row=i, column=1, value=lbl)
        cv = ws.cell(row=i, column=2, value=val)
        cl.font = Font(name=fn, size=9, bold=True)
        cv.font = Font(name=fn, size=9)
        cl.fill = hfl("EEF3FA")
        for c in (cl,cv):
            c.border = tb(); c.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 18


# ── 사업목록 시트 ────────────────────────────────────────────
def build_projects(ws, projects, cfg):
    sc  = cfg.get("sheet_projects",{})
    sty = cfg.get("style",{})
    fn  = sty.get("font_name","맑은 고딕")
    fs  = sty.get("font_size",9)
    hbg = sty.get("header_bg","1F3864")
    hfg = sty.get("header_fg","FFFFFF")
    se  = sty.get("stripe_even","F2F7FF")
    so  = sty.get("stripe_odd","FFFFFF")

    ws.title = sc.get("name","사업목록")
    cols    = list(sc.get("columns",{}).items())   # [(field, header)]
    widths  = sc.get("column_widths",{})
    nfmts   = sc.get("number_formats",{})
    cfmts   = sc.get("conditional_formats",{})
    ncols   = len(cols)

    # 타이틀
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f"AI 재정사업 총괄 — 사업목록  ({len(projects)}건)")
    c.font = Font(name=fn, size=12, bold=True, color="FFFFFF")
    c.fill = hfl(hbg); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value="※ 예산 단위: 백만원  |  R&D·정보화: O = 해당  |  읽기 전용")
    c.font = Font(name=fn, size=8, color="666666", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15

    # 헤더
    cf_col = None
    for ci,(field,header) in enumerate(cols,1):
        c = ws.cell(row=3, column=ci, value=header)
        c.font = Font(name=fn, size=fs, bold=True, color=hfg)
        c.fill = hfl(hbg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = hb()
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(header,12)
        if field in cfmts: cf_col = ci
    ws.row_dimensions[3].height = 32

    # 데이터
    for ri, p in enumerate(projects, 4):
        stripe = se if ri%2==0 else so
        for ci,(field,header) in enumerate(cols,1):
            raw = gn(p,field) if "." in field else p.get(field)
            val = fv(raw)
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(name=fn, size=fs, bold=(field=="department"))
            c.fill   = hfl(stripe); c.border = tb()
            c.alignment = Alignment(vertical="center", wrap_text=False)
            if field in nfmts and isinstance(val,(int,float)):
                c.number_format = nfmts[field]
                c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[ri].height = sc.get("row_height",18)

    # 조건부 서식 (증감률)
    if cf_col:
        for field, cf in cfmts.items():
            cl = get_column_letter(cf_col)
            rng = f"{cl}4:{cl}{3+len(projects)}"
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="greaterThan",formula=["0"],fill=hfl(cf.get("positive","C6EFCE"))))
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="lessThan",formula=["0"],fill=hfl(cf.get("negative","FFC7CE"))))

    if sc.get("auto_filter",True):
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"
    ws.freeze_panes = sc.get("freeze_panes","A4")


# ── 내역사업 시트 ────────────────────────────────────────────
def build_sub(ws, projects, cfg):
    sc  = cfg.get("sheet_sub_projects",{})
    sty = cfg.get("style",{})
    fn  = sty.get("font_name","맑은 고딕")
    fs  = sty.get("font_size",9)
    hbg = sty.get("header_bg","1F3864")
    hfg = sty.get("header_fg","FFFFFF")
    se  = sty.get("stripe_even","F2F7FF")
    so  = sty.get("stripe_odd","FFFFFF")

    ws.title = sc.get("name","내역사업")
    cols    = list(sc.get("columns",{}).items())
    widths  = sc.get("column_widths",{})
    nfmts   = sc.get("number_formats",{})
    ncols   = len(cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="AI 재정사업 총괄 — 내역사업")
    c.font = Font(name=fn, size=12, bold=True, color="FFFFFF")
    c.fill = hfl(hbg); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci,(field,header) in enumerate(cols,1):
        c = ws.cell(row=2, column=ci, value=header)
        c.font = Font(name=fn, size=fs, bold=True, color=hfg)
        c.fill = hfl(hbg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = hb()
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(header,14)
    ws.row_dimensions[2].height = sc.get("header_row_height",28)

    ri = 3
    for p in projects:
        for sub in p.get("sub_projects",[]):
            sub["parent_department"] = p.get("department","")
            sub["parent_code"]       = p.get("code","")
            sub["parent_name"]       = p.get("project_name","")
            stripe = se if ri%2==0 else so
            for ci,(field,_) in enumerate(cols,1):
                raw = sub.get(field)
                val = fv(raw)
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name=fn, size=fs, bold=(field=="parent_department"))
                c.fill = hfl(stripe); c.border = tb()
                c.alignment = Alignment(vertical="center")
                if field in nfmts and isinstance(val,(int,float)):
                    c.number_format = nfmts[field]
                    c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[ri].height = sc.get("row_height",18)
            ri += 1

    if sc.get("auto_filter",True):
        ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"
    ws.freeze_panes = sc.get("freeze_panes","A3")


# ── 연혁 시트 ────────────────────────────────────────────────
def build_history(ws, projects, cfg):
    sc  = cfg.get("sheet_history",{})
    if not sc.get("enabled", False): return
    fn  = cfg.get("style",{}).get("font_name","맑은 고딕")
    fs  = cfg.get("style",{}).get("font_size",9)
    hbg = cfg.get("style",{}).get("header_bg","1F3864")
    hfg = cfg.get("style",{}).get("header_fg","FFFFFF")
    se  = cfg.get("style",{}).get("stripe_even","F2F7FF")
    so  = cfg.get("style",{}).get("stripe_odd","FFFFFF")

    ws.title = sc.get("name","사업연혁")
    for col,w in zip("ABCD",[14,30,8,50]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "AI 재정사업 총괄 — 사업연혁"
    c.font = Font(name=fn,size=12,bold=True,color="FFFFFF")
    c.fill = hfl(hbg); c.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=28

    for ci,h in enumerate(["부처명","사업명","연도","내용"],1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=Font(name=fn,size=fs,bold=True,color=hfg); c.fill=hfl(hbg)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=hb()
    ws.row_dimensions[2].height=28

    ri=3
    for p in projects:
        for h in p.get("history",[]):
            stripe=se if ri%2==0 else so
            for ci,val in enumerate([p.get("department",""),p.get("project_name",""),
                                      h.get("year",""),h.get("description","")],1):
                c=ws.cell(row=ri,column=ci,value=fv(val))
                c.font=Font(name=fn,size=fs)
                c.fill=hfl(stripe); c.border=tb()
                c.alignment=Alignment(vertical="top",wrap_text=(ci==4))
            ws.row_dimensions[ri].height=18; ri+=1

    ws.freeze_panes="A3"


# ── 시트 보호 ─────────────────────────────────────────────────
def protect(ws, pw=""):
    ws.protection.sheet=True
    ws.protection.password=pw
    ws.protection.enable()


# ── 메인 ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="merged.json → 총괄 XLSX v2")
    parser.add_argument("input", nargs="?", default=str(path_config.MERGED_JSON_PATH))
    parser.add_argument("--config", default=str(ROOT/"config"/"config_export.yaml"))
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    cfg_path = Path(args.config)
    cfg = load_cfg(cfg_path) if cfg_path.exists() else {}

    # _years 기반으로 config_export.yaml의 연도 레이블 동적 교체
    Y = get_years()
    # sproject columns의 연도 레이블 교체
    sp_cols = cfg.get("sheet_projects", {}).get("columns", {})
    label_map = {
        f"budget.{Y['settlement']}_settlement": Y["label_settlement"],
        f"budget.{Y['original']}_original":    Y["label_original"],
        f"budget.{Y['original']}_supplementary": Y["label_supplementary"],
        f"budget.{Y['budget']}_request":       Y["label_request"],
        f"budget.{Y['budget']}_budget":        Y["label_budget"],
    }
    for fld, new_label in label_map.items():
        if fld in sp_cols:
            old_label = sp_cols[fld]
            sp_cols[fld] = new_label
            # column_widths도 키 교체
            cw = cfg.get("sheet_projects", {}).get("column_widths", {})
            if old_label in cw:
                cw[new_label] = cw.pop(old_label)
            # number_formats도 키 교체
            nf = cfg.get("sheet_projects", {}).get("number_formats", {})
            # (필드명 기준이라 변경 불필요)

    # sub_projects columns 연도 레이블 교체
    sub_cols = cfg.get("sheet_sub_projects", {}).get("columns", {})
    sub_label_map = {
        "budget_2024": Y["label_sub"][0],
        "budget_2025": Y["label_sub"][1],
        "budget_2026": Y["label_sub"][2],
    }
    for fld, new_label in sub_label_map.items():
        if fld in sub_cols:
            old_label = sub_cols[fld]
            sub_cols[fld] = new_label
            scw = cfg.get("sheet_sub_projects", {}).get("column_widths", {})
            if old_label in scw:
                scw[new_label] = scw.pop(old_label)

    with open(args.input, encoding="utf-8") as f: data = json.load(f)
    projects = data.get("projects",data) if isinstance(data,dict) else data
    metadata = data.get("metadata",{})   if isinstance(data,dict) else {}
    print(f"▶ {len(projects)}건 로드")

    out_cfg = cfg.get("output",{})
    if args.out:
        out_path = Path(args.out)
    else:
        today   = datetime.date.today().strftime("%Y%m%d")
        prefix  = out_cfg.get("filename_prefix","총괄_AI재정사업")
        out_dir = path_config.OUTPUT_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{prefix}_{today}.xlsx"

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    pw = cfg.get("style",{}).get("protect_password","")
    do_protect = cfg.get("style",{}).get("protect_sheets",True)

    if cfg.get("sheet_summary",{}).get("enabled",True):
        ws=wb.create_sheet(); build_summary(ws,metadata,cfg)
        if do_protect: protect(ws,pw)

    ws=wb.create_sheet(); build_projects(ws,projects,cfg)
    if do_protect: protect(ws,pw)

    ws=wb.create_sheet(); build_sub(ws,projects,cfg)
    if do_protect: protect(ws,pw)

    if cfg.get("sheet_history",{}).get("enabled",False):
        ws=wb.create_sheet(); build_history(ws,projects,cfg)
        if do_protect: protect(ws,pw)

    wb.save(out_path)
    print(f"✅ 완료: {out_path}")
    print(f"   시트: {[s.title for s in wb.worksheets]}")

if __name__=="__main__":
    main()
