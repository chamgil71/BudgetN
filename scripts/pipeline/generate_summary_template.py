"""
generate_summary_template.py
convert.py가 바로 읽을 수 있는 총괄 XLSX 템플릿을 생성한다.
"""
import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "scripts" / "pipeline"))

from convert import ROOT, get_years, load_config


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=9)
THIN = Side(style="thin", color="D9E2F3")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BOX


def style_note(cell):
    cell.font = Font(name="맑은 고딕", size=9, italic=True, color="666666")
    cell.alignment = Alignment(vertical="center")


def style_body(cell):
    cell.font = BODY_FONT
    cell.alignment = Alignment(vertical="center")
    cell.border = BOX


def ordered_headers(mapping):
    headers = []
    seen = set()
    for header in mapping.keys():
        normalized = str(header).replace("★", "").strip()
        if normalized in seen:
            continue
        seen.add(normalized)
        headers.append(normalized)
    return headers


def create_sheet(ws, title, headers, note):
    ws.title = title
    ws.cell(row=1, column=1, value=note)
    style_note(ws.cell(row=1, column=1))

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=idx, value=header)
        style_header(cell)
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(12, min(len(str(header)) + 4, 30))

        body = ws.cell(row=3, column=idx, value=None)
        style_body(body)

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 24


def build_workbook(cfg):
    xlsx = cfg["xlsx"]
    Y = get_years(cfg)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    project_headers = ordered_headers(xlsx.get("column_mapping", {}))
    sub_headers = ordered_headers(xlsx.get("sub_projects_mapping", {}))
    manager_headers = ordered_headers(xlsx.get("managers_mapping", {}))
    history_headers = ordered_headers(xlsx.get("history_mapping", {}))
    yearly_headers = ["상위사업코드", *[str(y) for y in Y["sub_years"]]]

    create_sheet(
        wb.create_sheet(),
        xlsx["data_sheet"],
        project_headers,
        "2행은 헤더, 3행부터 데이터 입력. 입력 후 실행: python scripts/pipeline/excel_manager.py import --type summary --file <작성한파일.xlsx>",
    )
    create_sheet(
        wb.create_sheet(),
        xlsx["sub_projects_sheet"],
        sub_headers,
        "상위사업코드는 사업목록 시트의 사업코드와 일치해야 합니다. 내역사업은 여러 행으로 입력할 수 있습니다.",
    )
    create_sheet(
        wb.create_sheet(),
        xlsx["managers_sheet"],
        manager_headers,
        "담당자 시트는 선택 입력입니다. 상위사업코드 기준으로 사업과 연결됩니다.",
    )
    create_sheet(
        wb.create_sheet(),
        xlsx["history_sheet"],
        history_headers,
        "사업연혁 시트는 선택 입력입니다. 연도와 내용을 행 단위로 추가합니다.",
    )
    create_sheet(
        wb.create_sheet(),
        xlsx["yearly_budget_sheet"],
        yearly_headers,
        f"연도별예산 시트는 선택 입력입니다. 기본 연도 열은 {', '.join(map(str, Y['sub_years']))}입니다.",
    )
    return wb


def main():
    parser = argparse.ArgumentParser(description="총괄 XLSX import 템플릿 생성")
    parser.add_argument("--out", default=None, help="출력 경로. 기본값은 config.yaml의 xlsx.template_file")
    args = parser.parse_args()

    cfg = load_config()
    out = Path(args.out) if args.out else ROOT / cfg["xlsx"].get("template_file", "template_project.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook(cfg)
    wb.save(out)
    print(f"Template created: {out}")
    print(f"   시트: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
