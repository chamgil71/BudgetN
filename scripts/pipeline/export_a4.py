"""
export_a4.py v2.0  —  merged.json → A4 요약표 XLSX
- 전체 필드 정합 (budget_db.json 기준)
- Named Range 태깅 (역방향 추출용)
- 섹션별 config on/off
실행:
  python scripts/export_a4.py
  python scripts/export_a4.py --id 1134-309
  python scripts/export_a4.py --dept 과학기술정보통신부
  python scripts/export_a4.py --limit 10
"""
import sys, os, re, json, yaml, datetime, argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("pip install openpyxl --break-system-packages"); sys.exit(1)

ROOT = Path(__file__).parent.parent.parent
import sys as _sys; _sys.path.insert(0, str(ROOT/"scripts"))
from _years import get_years as _get_years_module

def get_years(cfg=None):
    return _get_years_module(cfg if cfg else ROOT/"config"/"config.yaml")
ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
FN = "맑은 고딕"


# ── 헬퍼 ─────────────────────────────────────────────────────
def load_cfg(p):
    with open(p, encoding="utf-8") as f: return yaml.safe_load(f)

def gn(obj, path):
    if not isinstance(obj, dict): return None
    parts = path.split(".", 1)
    if len(parts) == 1: return obj.get(path)
    s = obj.get(parts[0])
    return gn(s, parts[1]) if isinstance(s, dict) else None

def clean(v):
    if v is None: return ""
    if isinstance(v, bool): return "O" if v else "X"
    if isinstance(v, list): return ", ".join(ILLEGAL.sub("", str(x)) for x in v)
    if isinstance(v, str): return ILLEGAL.sub("", v)
    return v

def hfl(c): return PatternFill("solid", fgColor=c.lstrip("#"))
def bd(c="CCCCCC"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def hbd(bc="2E74B5"):
    s = Side(style="thin", color="CCCCCC")
    b = Side(style="medium", color=bc)
    return Border(left=s, right=s, top=s, bottom=b)

def nr_key(sname, field):
    """Named Range 키 생성: _시트prefix_field__path"""
    prefix = re.sub(r'[^A-Za-z0-9]', '_', sname)[:8]
    fkey   = re.sub(r'[^A-Za-z0-9_]', '_', field.replace(".", "__"))
    return f"_{prefix}_{fkey}"

def register_nr(wb, name, ref):
    """Named Range 등록 (중복 무시)"""
    try:
        dn = DefinedName(name=name, attr_text=ref)
        wb.defined_names[name] = dn
    except Exception:
        pass


# ── 셀 스타일 적용 헬퍼 ──────────────────────────────────────
def title_cell(ws, row, col, ncols, text, bg="1F3864", size=11, height=22):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ncols-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FN, size=size, bold=True, color="FFFFFF")
    c.fill = hfl(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    return c

def section_cell(ws, row, ncols, text, bg="D6E4F0", height=16):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FN, size=9, bold=True, color="1F3864")
    c.fill = hfl(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(
        top=Side(style="medium", color="2E74B5"),
        bottom=Side(style="medium", color="2E74B5"),
    )
    ws.row_dimensions[row].height = height

def label_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FN, size=8, bold=True, color="2E74B5")
    c.fill = hfl("EEF4FB")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = bd()
    return c

def value_cell(ws, row, col, val, span=1, wrap=False, nfmt=None, bg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FN, size=9)
    c.fill = hfl(bg)
    c.alignment = Alignment(horizontal="left", vertical="top" if wrap else "center",
                            wrap_text=wrap)
    c.border = bd()
    if nfmt: c.number_format = nfmt
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def table_header(ws, row, cols_widths, bg="2E74B5"):
    """표 헤더 행 생성. cols_widths = [(헤더텍스트, 너비or None), ...]"""
    for ci, (h, _) in enumerate(cols_widths, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name=FN, size=8, bold=True, color="FFFFFF")
        c.fill = hfl(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bd()
    ws.row_dimensions[row].height = 16


# ════════════════════════════════════════════════════════════
# 시트 빌더
# ════════════════════════════════════════════════════════════
def build_sheet(wb, ws, p, cfg, Y):
    sec  = cfg.get("sections", {})
    sty  = cfg.get("style", {})
    sname = ws.title
    NC = 5   # 열 수 (A~E)
    named = {}  # {field: "셀참조"}

    # 열 너비
    for col, w in zip("ABCDE", [14, 28, 12, 12, 12]):
        ws.column_dimensions[col].width = w

    row = 1

    # ── 헤더 ────────────────────────────────────────────────
    if sec.get("header", True):
        # 행1: 부처 + 코드
        dept = clean(p.get("department",""))
        code = clean(p.get("code",""))
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=dept)
        c.font = Font(name=FN, size=10, bold=True, color="FFFFFF")
        c.fill = hfl("1F3864"); c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        named["department"] = f"'{sname}'!$A${row}"

        ws.merge_cells(f"D{row}:E{row}")
        c2 = ws.cell(row=row, column=4, value=f"코드: {code}")
        c2.font = Font(name=FN, size=9, color="FFFFFF")
        c2.fill = hfl("1F3864"); c2.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        named["code"] = f"'{sname}'!$D${row}"
        ws.row_dimensions[row].height = 20; row += 1

        # 행2: 사업명
        pname = clean(p.get("project_name",""))
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1, value=pname)
        c.font = Font(name=FN, size=13, bold=True, color="FFFFFF")
        c.fill = hfl("1F3864"); c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        named["project_name"] = f"'{sname}'!$A${row}"
        ws.row_dimensions[row].height = 26; row += 1

    # ── 기본 정보 ────────────────────────────────────────────
    if sec.get("basic_info", True):
        section_cell(ws, row, NC, "■ 기본 정보"); row += 1

        basic_pairs = [
            [("회계유형","account_type"), ("분야","field")],
            [("부문","sector"),          ("상태","status")],
            [("지원유형","support_type"),("시행기관","implementing_agency")],
            [("R&D 여부","is_rnd"),      ("정보화","is_informatization")],
            [("보조율","subsidy_rate"),  ("융자율","loan_rate")],
        ]
        for pairs in basic_pairs:
            for (lbl, fld), (lc, vc) in zip(pairs, [(1,2),(3,4)]):
                label_cell(ws, row, lc, lbl)
                val = clean(p.get(fld,""))
                c = value_cell(ws, row, vc, val)
                named[fld] = f"'{sname}'!${get_column_letter(vc)}${row}"
            ws.cell(row=row, column=5).fill = hfl("FFFFFF")
            ws.cell(row=row, column=5).border = bd()
            ws.row_dimensions[row].height = 17; row += 1

        # 프로그램 계층
        for lbl, fld in [("프로그램","program"), ("단위사업","unit_project"), ("세부사업","detail_project")]:
            obj = p.get(fld) or {}
            code_v = clean(obj.get("code",""))
            name_v = clean(obj.get("name",""))
            label_cell(ws, row, 1, lbl)
            c = ws.cell(row=row, column=2, value=f"[{code_v}] {name_v}" if code_v else name_v)
            c.font = Font(name=FN, size=9); c.fill = hfl("FFFFFF")
            c.alignment = Alignment(vertical="center"); c.border = bd()
            ws.merge_cells(f"B{row}:E{row}")
            named[f"{fld}.name"] = f"'{sname}'!$B${row}"
            ws.row_dimensions[row].height = 17; row += 1

    # ── 사업 개요 ─────────────────────────────────────────────
    overview = p.get("overview") or {}
    if sec.get("overview", True) and overview:
        section_cell(ws, row, NC, "■ 사업 개요"); row += 1
        for k, v in overview.items():
            label_cell(ws, row, 1, k)
            value_cell(ws, row, 2, clean(v), span=4, wrap=True)
            ws.row_dimensions[row].height = 17; row += 1

    # ── 예산 현황 ─────────────────────────────────────────────
    if sec.get("budget_table", True):
        section_cell(ws, row, NC,
            f"■ 예산 현황 (단위: 백만원)  ┃  증감액·증감률 자동계산"); row += 1

        bhdrs = [
            (f"{Y['settlement']} 결산",  "budget.2024_settlement"),
            (f"{Y['original']} 본예산",  "budget.2025_original"),
            (f"{Y['budget']} 확정",      "budget.2026_budget"),
            ("증감액 ▶자동",             "budget.change_amount"),
            ("증감률(%) ▶자동",          "budget.change_rate"),
        ]
        for ci,(h,_) in enumerate(bhdrs,1):
            c=ws.cell(row=row,column=ci,value=h)
            c.font=Font(name=FN,size=8,bold=True,color="FFFFFF")
            c.fill=hfl("2E74B5")
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bd()
        ws.row_dimensions[row].height=17; row+=1

        b26=get_column_letter(3); b25=get_column_letter(2)
        for ci,(h,fld) in enumerate(bhdrs,1):
            c=ws.cell(row=row,column=ci)
            is_auto = "자동" in h
            if is_auto:
                if "증감액" in h:
                    c.value=f"=IF(AND({b26}{row}<>\"\",{b25}{row}<>\"\"),{b26}{row}-{b25}{row},\"\")"
                    c.number_format="#,##0;(#,##0)"
                else:
                    c.value=f"=IF(AND({b26}{row}<>\"\",{b25}{row}<>\"\",{b25}{row}<>0),({b26}{row}-{b25}{row})/{b25}{row}*100,\"\")"
                    c.number_format="0.0;(0.0)"
                c.fill=hfl("F5F5F5"); c.font=Font(name=FN,size=9,color="555555",italic=True)
            else:
                raw=gn(p,fld)
                c.value=raw if isinstance(raw,(int,float)) and raw is not None else None
                c.number_format="#,##0"
                c.fill=hfl("FFF8E1" if fld.endswith("budget") else "FFFFFF")
                c.font=Font(name=FN,size=9)
                named[fld]=f"'{sname}'!${get_column_letter(ci)}${row}"
            c.alignment=Alignment(horizontal="right",vertical="center"); c.border=bd()
        ws.row_dimensions[row].height=17; row+=1

        # 추경·요구 (별도 행)
        if sec.get("budget_detail", True):
            for lbl,fld in [(f"{Y['supplementary']} 추경","budget.2025_supplementary"),
                             (f"{Y['request']} 요구","budget.2026_request")]:
                label_cell(ws,row,1,lbl)
                raw=gn(p,fld)
                c=value_cell(ws,row,2,raw,span=4)
                c.number_format="#,##0"; c.alignment=Alignment(horizontal="right",vertical="center")
                named[fld]=f"'{sname}'!$B${row}"
                ws.row_dimensions[row].height=16; row+=1

        # 연도별예산 (yearly_budgets)
        yb = p.get("yearly_budgets") or {}
        if sec.get("yearly_budgets", True) and len(yb) > 1:
            yb_sorted = sorted(yb.items(), key=lambda x: x[0])
            label_cell(ws,row,1,"연도별 이력")
            yr_text = "  ".join(f"{yr}: {int(v):,}" for yr,v in yb_sorted if v is not None)
            c=value_cell(ws,row,2,yr_text,span=4)
            c.font=Font(name=FN,size=8,color="555555")
            ws.row_dimensions[row].height=16; row+=1

    # ── 사업 기간 / 총사업비 ──────────────────────────────────
    if sec.get("period_cost", True):
        section_cell(ws, row, NC, "■ 사업기간 및 총사업비"); row += 1
        pp = p.get("project_period") or {}
        tc = p.get("total_cost") or {}
        pairs = [
            [("사업시작연도","project_period.start_year"),("사업종료연도","project_period.end_year")],
            [("사업기간(원문)","project_period.raw"),    ("사업기간(년)","project_period.duration")],
            [("총사업비","total_cost.total"),            ("국고","total_cost.government")],
        ]
        for pair in pairs:
            for (lbl,fld),(lc,vc) in zip(pair,[(1,2),(3,4)]):
                label_cell(ws,row,lc,lbl)
                raw=gn(p,fld)
                c=value_cell(ws,row,vc,clean(raw) if not isinstance(raw,(int,float)) else raw)
                if isinstance(raw,(int,float)) and raw is not None:
                    c.number_format="#,##0"; c.alignment=Alignment(horizontal="right",vertical="center")
                named[fld]=f"'{sname}'!${get_column_letter(vc)}${row}"
            ws.cell(row=row,column=5).fill=hfl("FFFFFF"); ws.cell(row=row,column=5).border=bd()
            ws.row_dimensions[row].height=17; row+=1

    # ── AI 분류 ───────────────────────────────────────────────
    if sec.get("ai_classification", True):
        section_cell(ws, row, NC, "■ AI 분류"); row += 1
        for lbl,fld in [("AI 도메인","ai_domains"),("AI 기술유형","ai_tech"),
                         ("R&D 단계","rnd_stage"),("키워드","keywords")]:
            val=clean(p.get(fld,"") or "")
            if not val: continue
            label_cell(ws,row,1,lbl)
            c=value_cell(ws,row,2,val,span=4,wrap=True)
            named[fld]=f"'{sname}'!$B${row}"
            ws.row_dimensions[row].height=17; row+=1

    # ── 내역사업 ──────────────────────────────────────────────
    subs = p.get("sub_projects",[])
    if sec.get("sub_projects", True):
        section_cell(ws, row, NC,
            f"■ 내역사업 ({len(subs)}건)  ┃  빈 행 없이 연속 입력"); row += 1

        SUB = Y.get("sub_years",[2024,2025,2026])
        # 헤더: A~B 병합=내역사업명, C=연도1, D=연도2, E=연도3
        ws.merge_cells(f"A{row}:B{row}")
        c=ws.cell(row=row,column=1,value="내역사업명")
        c.font=Font(name=FN,size=8,bold=True,color="FFFFFF")
        c.fill=hfl("2E74B5"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bd()
        for ci2,yr in zip([3,4,5], SUB):
            c=ws.cell(row=row,column=ci2,value=str(yr))
            c.font=Font(name=FN,size=8,bold=True,color="FFFFFF")
            c.fill=hfl("2E74B5"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bd()
        # B열 헤더 셀 스타일 (병합이지만 스타일 적용)
        ws.cell(row=row,column=2).fill=hfl("2E74B5"); ws.cell(row=row,column=2).border=bd()
        ws.row_dimensions[row].height=16; row+=1

        display_subs = subs if subs else [{"name":"(내역사업 없음)"}]
        for si,sub in enumerate(display_subs):
            bg="F2F7FF" if si%2==0 else "FFFFFF"
            # A~B 병합=사업명, C=연도1예산, D=연도2예산, E=연도3예산
            ws.merge_cells(f"A{row}:B{row}")
            c=ws.cell(row=row,column=1,value=clean(sub.get("name","")))
            c.font=Font(name=FN,size=9); c.fill=hfl(bg); c.border=bd()
            c.alignment=Alignment(vertical="center")
            ws.cell(row=row,column=2).fill=hfl(bg); ws.cell(row=row,column=2).border=bd()
            for ci2,fk in zip([3,4,5],["budget_2024","budget_2025","budget_2026"]):
                v=sub.get(fk)
                c=ws.cell(row=row,column=ci2,value=v)
                c.font=Font(name=FN,size=9); c.fill=hfl(bg); c.border=bd()
                if v is not None: c.number_format="#,##0"
                c.alignment=Alignment(horizontal="right",vertical="center")
            ws.row_dimensions[row].height=17; row+=1

    # ── 사업관리자 ────────────────────────────────────────────
    mgrs = p.get("project_managers",[])
    if sec.get("project_managers", True) and mgrs:
        section_cell(ws, row, NC, f"■ 사업관리자 ({len(mgrs)}건)"); row += 1
        for ci,h in enumerate(["내역사업명","소관부서","시행기관","담당자","연락처"],1):
            c=ws.cell(row=row,column=ci,value=h)
            c.font=Font(name=FN,size=8,bold=True,color="FFFFFF")
            c.fill=hfl("2E74B5"); c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=bd()
        ws.row_dimensions[row].height=16; row+=1

        for mi,mgr in enumerate(mgrs):
            bg="F2F7FF" if mi%2==0 else "FFFFFF"
            for ci,fk in enumerate(["sub_project","managing_dept","implementing_agency","manager","phone"],1):
                c=ws.cell(row=row,column=ci,value=clean(mgr.get(fk,"")))
                c.font=Font(name=FN,size=8); c.fill=hfl(bg); c.border=bd()
                c.alignment=Alignment(vertical="center")
            ws.row_dimensions[row].height=16; row+=1

    # ── 성과지표 KPI ──────────────────────────────────────────
    kpi_list = p.get("kpi",[])
    if sec.get("kpi", True) and kpi_list:
        section_cell(ws, row, NC, f"■ 성과지표 ({len(kpi_list)}건)"); row += 1
        for ki, kpi in enumerate(kpi_list):
            bg = "F2F7FF" if ki%2==0 else "FFFFFF"
            label_cell(ws,row,1,"지표명")
            c=value_cell(ws,row,2,clean(kpi.get("name","")),span=3)
            ws.cell(row=row,column=5,value=f"가중치:{kpi.get('weight','')}")
            ws.cell(row=row,column=5).font=Font(name=FN,size=8,color="555555")
            ws.cell(row=row,column=5).border=bd()
            ws.row_dimensions[row].height=17; row+=1

            # 목표/실적 (연도별)
            targets=kpi.get("targets",{}); actuals=kpi.get("actuals",{})
            yrs=sorted(set(list(targets.keys())+list(actuals.keys())))
            if yrs:
                label_cell(ws,row,1,"연도")
                for ci2,yr in enumerate(yrs[:4],2):
                    c=ws.cell(row=row,column=ci2,value=yr)
                    c.font=Font(name=FN,size=8,bold=True); c.fill=hfl("EEF4FB")
                    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bd()
                ws.row_dimensions[row].height=15; row+=1

                label_cell(ws,row,1,"목표")
                for ci2,yr in enumerate(yrs[:4],2):
                    value_cell(ws,row,ci2,targets.get(yr,""))
                ws.row_dimensions[row].height=15; row+=1

                label_cell(ws,row,1,"실적")
                for ci2,yr in enumerate(yrs[:4],2):
                    value_cell(ws,row,ci2,actuals.get(yr,""))
                ws.row_dimensions[row].height=15; row+=1

            if kpi.get("measurement_method"):
                label_cell(ws,row,1,"측정방법")
                value_cell(ws,row,2,clean(kpi.get("measurement_method","")),span=4,wrap=True)
                ws.row_dimensions[row].height=15; row+=1

    # ── 사업연혁 ──────────────────────────────────────────────
    hist = p.get("history",[])
    if sec.get("history", True) and hist:
        section_cell(ws, row, NC, f"■ 사업연혁 ({len(hist)}건)"); row += 1
        for ci,h in enumerate(["연도","주요 내용"],1):
            c=ws.cell(row=row,column=ci,value=h)
            c.font=Font(name=FN,size=8,bold=True,color="FFFFFF")
            c.fill=hfl("2E74B5"); c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=bd()
            if ci==2: ws.merge_cells(f"B{row}:E{row}")
        ws.row_dimensions[row].height=15; row+=1

        for hi,h in enumerate(hist):
            bg="F2F7FF" if hi%2==0 else "FFFFFF"
            c=ws.cell(row=row,column=1,value=h.get("year",""))
            c.font=Font(name=FN,size=8); c.fill=hfl(bg); c.border=bd()
            c.alignment=Alignment(horizontal="center",vertical="top")
            desc=clean(h.get("description",""))
            c2=ws.cell(row=row,column=2,value=desc)
            c2.font=Font(name=FN,size=8); c2.fill=hfl(bg); c2.border=bd()
            c2.alignment=Alignment(vertical="top",wrap_text=True)
            ws.merge_cells(f"B{row}:E{row}")
            lines=max(1,min(6,desc.count('\n')+1+len(desc)//60))
            ws.row_dimensions[row].height=max(16,lines*13); row+=1

    # ── 효과성 ────────────────────────────────────────────────
    eff = p.get("effectiveness","")
    if sec.get("effectiveness", True) and eff and eff != "(작성예정)":
        section_cell(ws, row, NC, "■ 효과성"); row += 1
        txt=clean(eff)[:1000]
        c=ws.cell(row=row,column=1,value=txt)
        c.font=Font(name=FN,size=9); c.fill=hfl("FFFFFF"); c.border=bd()
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        ws.merge_cells(f"A{row}:E{row}")
        named["effectiveness"]=f"'{sname}'!$A${row}"
        lines=max(3,min(12,txt.count('\n')+2+len(txt)//80))
        ws.row_dimensions[row].height=lines*13; row+=1

    # ── 집행 정보 ─────────────────────────────────────────────
    ed = p.get("execution_detail") or {}
    bc = p.get("budget_calculation")
    if sec.get("execution", True) and (any(ed.values()) or bc):
        section_cell(ws, row, NC, "■ 집행 정보"); row += 1
        for lbl,fld,key in [
            ("집행방법","execution_detail.method","method"),
            ("사업수혜자","execution_detail.recipients","recipients"),
        ]:
            val=ed.get(key,"") or ""
            if not val: continue
            label_cell(ws,row,1,lbl)
            value_cell(ws,row,2,clean(val),span=4,wrap=True)
            named[fld]=f"'{sname}'!$B${row}"
            ws.row_dimensions[row].height=17; row+=1
        if bc:
            label_cell(ws,row,1,"예산산출근거")
            bc_text = "\n".join(bc) if isinstance(bc,list) else clean(bc)
            c=value_cell(ws,row,2,bc_text,span=4,wrap=True)
            named["budget_calculation"]=f"'{sname}'!$B${row}"
            lines=max(2,min(8,bc_text.count('\n')+2))
            ws.row_dimensions[row].height=lines*13; row+=1

    # ── 사업목적 ──────────────────────────────────────────────
    if sec.get("purpose", True) and p.get("purpose"):
        section_cell(ws, row, NC, "■ 사업목적  (Alt+Enter 줄바꿈 가능)"); row += 1
        txt=clean(p.get("purpose",""))[:2000]
        c=ws.cell(row=row,column=1,value=txt)
        c.font=Font(name=FN,size=9); c.fill=hfl("FFFFFF"); c.border=bd()
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        ws.merge_cells(f"A{row}:E{row}")
        named["purpose"]=f"'{sname}'!$A${row}"
        lines=max(3,min(15,txt.count('\n')+2+len(txt)//80))
        ws.row_dimensions[row].height=lines*13; row+=1

    # ── 사업내용 ──────────────────────────────────────────────
    if sec.get("description", True) and p.get("description"):
        section_cell(ws, row, NC, "■ 사업내용"); row += 1
        txt=clean(p.get("description",""))[:2000]
        c=ws.cell(row=row,column=1,value=txt)
        c.font=Font(name=FN,size=9); c.fill=hfl("FFFFFF"); c.border=bd()
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        ws.merge_cells(f"A{row}:E{row}")
        named["description"]=f"'{sname}'!$A${row}"
        lines=max(3,min(15,txt.count('\n')+2+len(txt)//80))
        ws.row_dimensions[row].height=lines*13; row+=1

    # ── 법적근거 ──────────────────────────────────────────────
    if sec.get("legal_basis", False) and p.get("legal_basis"):
        section_cell(ws, row, NC, "■ 법적근거"); row += 1
        txt=clean(p.get("legal_basis",""))[:1500]
        c=ws.cell(row=row,column=1,value=txt)
        c.font=Font(name=FN,size=8); c.fill=hfl("FFFFFF"); c.border=bd()
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        ws.merge_cells(f"A{row}:E{row}")
        named["legal_basis"]=f"'{sname}'!$A${row}"
        lines=max(2,min(10,txt.count('\n')+2+len(txt)//80))
        ws.row_dimensions[row].height=lines*13; row+=1

    # ── 푸터 ─────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    c=ws.cell(row=row,column=1,
        value=f"출력일: {datetime.date.today()}  |  KAIB 파이프라인 자동 생성  |  사업코드: {p.get('code','')}")
    c.font=Font(name=FN,size=7,color="AAAAAA",italic=True)
    c.alignment=Alignment(horizontal="right",vertical="center")
    ws.row_dimensions[row].height=13

    # ── Named Range 등록 ─────────────────────────────────────
    if cfg.get("enable_named_ranges", True):
        for fld, ref in named.items():
            nk = nr_key(sname, fld)
            register_nr(wb, nk, ref)

    # ── 인쇄 설정 ────────────────────────────────────────────
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.59, right=0.39, top=0.75, bottom=0.75,
        header=0.31, footer=0.31
    )


# ── 메인 ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="merged.json → A4 요약표 v2")
    parser.add_argument("input", nargs="?", default=str(ROOT/"output"/"merged.json"))
    parser.add_argument("--config", default=str(ROOT/"config"/"config_a4.yaml"))
    parser.add_argument("--out",    default=None)
    parser.add_argument("--id",     default=None)
    parser.add_argument("--dept",   default=None)
    parser.add_argument("--status", default=None)
    parser.add_argument("--rnd",         action="store_true")
    parser.add_argument("--limit",       type=int, default=None)
    parser.add_argument("--split-by-dept", action="store_true",
        help="부처별로 파일 분리 생성 (시트 분리 아닌 파일 분리)")
    args = parser.parse_args()

    cfg = load_cfg(Path(args.config)) if Path(args.config).exists() else {}
    Y   = get_years(cfg)

    with open(args.input, encoding="utf-8") as f: data = json.load(f)
    projects = data.get("projects", data) if isinstance(data,dict) else data

    # 필터
    filt = cfg.get("filter",{})
    filtered = []
    for p in projects:
        if args.id     and p.get("code") != args.id: continue
        if args.dept   and p.get("department") != args.dept: continue
        if args.status and p.get("status") != args.status: continue
        if args.rnd    and not p.get("is_rnd"): continue
        if filt.get("department") and p.get("department") != filt["department"]: continue
        if filt.get("status")     and p.get("status") != filt["status"]: continue
        if filt.get("is_rnd") is not None and p.get("is_rnd") != filt["is_rnd"]: continue
        if filt.get("min_budget"):
            b = (p.get("budget") or {}).get("2026_budget") or 0
            if b < filt["min_budget"]: continue
        filtered.append(p)

    if args.limit: filtered = filtered[:args.limit]
    if not filtered: print("⚠️  출력할 사업 없음"); return

    print(f"▶ {len(filtered)}건")

    # 출력 경로
    out_cfg = cfg.get("output",{})
    if args.out:
        out_path = Path(args.out)
    else:
        today  = datetime.date.today().strftime("%Y%m%d")
        prefix = out_cfg.get("filename_prefix","A4요약_AI재정사업")
        suffix = f"_{args.dept}" if args.dept else ""
        out_dir = ROOT / out_cfg.get("dir","output")
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{prefix}{suffix}_{today}.xlsx"

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    for i, p in enumerate(filtered, 1):
        code  = re.sub(r'[\\/*?:\[\]]','_', str(p.get("code") or f"p{i}"))
        sname = code[:31]
        ws    = wb.create_sheet(title=sname)
        build_sheet(wb, ws, p, cfg, Y)
        if i % 50 == 0: print(f"  처리중: {i}/{len(filtered)}")

    if args.split_by_dept:
        # 부처별로 별도 파일 생성
        from collections import defaultdict
        import re as _re
        by_dept = defaultdict(list)
        for p in filtered: by_dept[p.get("department","기타")].append(p)
        for dept, projs in by_dept.items():
            safe = _re.sub(r'[\\/*?:<>|]','_',dept)
            dp_path = out_path.parent / f"A4요약_{safe}_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
            dwb = openpyxl.Workbook(); dwb.remove(dwb.active)
            for i2,p2 in enumerate(projs,1):
                code2 = re.sub(r'[\\/*?:\[\]]','_',str(p2.get("code") or f"p{i2}"))
                ws2 = dwb.create_sheet(title=code2[:31])
                build_sheet(dwb, ws2, p2, cfg, Y)
            dwb.save(dp_path)
            print(f"  ✅ {dept}: {len(projs)}건 → {dp_path.name}")
    else:
        wb.save(out_path)
        print(f"✅ 완료: {out_path}")
        print(f"   시트: {len(wb.worksheets)}개")

if __name__ == "__main__":
    main()
